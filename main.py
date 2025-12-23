import asyncio, aiohttp, openpyxl, random
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL, HEADERS = "https://www.olx.ua/api/v1/offers", {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36", "Accept": "application/json"}

async def fetch(s, url):
    async with s.get(url, headers=HEADERS) as r:
        return await r.json() if r.status == 200 else None

def extract(o):
    d = {"id": o.get("id"), "title": o.get("title"), "url": o.get("url"), "description": o.get("description"), "price": None, "currency": None, "negotiable": None, "city": o.get("location", {}).get("city", {}).get("name"), "region": o.get("location", {}).get("region", {}).get("name"), "created": o.get("created_time"), "photos": ", ".join([p["link"].replace(";s={width}x{height}", "") for p in o.get("photos", [])]), "seller": o.get("user", {}).get("name"), "business": o.get("business")}
    for p in o.get("params", []): 
        if p["key"] == "price": d.update({"price": p["value"].get("value"), "currency": p["value"].get("currency"), "negotiable": p["value"].get("negotiable")}); break
    return d

async def parse(s, url, ws, n):
    data = await fetch(s, url)
    if not data: return None
    offers = data.get("data", [])
    for o in offers: d = extract(o); ws.append([d["id"], d["title"], d["url"], d["description"], d["price"], d["currency"], d["negotiable"], d["city"], d["region"], d["created"], d["photos"], d["seller"], d["business"]])
    print(f"\033[92mСтраница {n}: {len(offers)} объявлений\033[0m", end="")
    return data.get("links", {}).get("next", {}).get("href")

async def main():
    wb, ws = openpyxl.Workbook(), None; ws = wb.active; ws.title = "OLX"
    ws.append(["ID", "Заголовок", "URL", "Описание", "Цена", "Валюта", "Торг", "Город", "Регион", "Дата", "Фото", "Продавец", "Бизнес"])
    ws.freeze_panes = "A2"
    for c in ws[1]: c.font, c.fill, c.alignment = Font(bold=True, color="FFFFFF"), PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"), Alignment(horizontal="center", wrap_text=True)
    for col in range(1, 14): ws.column_dimensions[get_column_letter(col)].width = 35
    async with aiohttp.ClientSession() as s:
        url, n = BASE_URL + "?offset=0", 1
        while url:
            url = await parse(s, url, ws, n)
            if n % 2 == 0 or not url:
                for row in range(2, ws.max_row + 1): ws.row_dimensions[row].height = 120
                wb.save("olx_offers.xlsx")
            if url: d = random.uniform(3, 6); print(f" → \033[93mЗадержка {d:.1f}с\033[0m"); await asyncio.sleep(d)
            else: print()
            n += 1

asyncio.run(main())